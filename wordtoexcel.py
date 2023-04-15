    from docx2python import docx2python
from openpyxl import Workbook

# Extract text from docx file
doc_result = docx2python('input.docx')
text = doc_result.text

# Split text into lines
lines = text.split('\n')

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set the title for the first column
ws.cell(row=1, column=1).value = 'Acceptance Criteria'

# Initialize row counter and cell text
row_counter = 2
cell_text = ''

# Iterate over lines and extract text between capital numerical bullets
for line in lines:
    if line.strip():
        if line[0].isdigit():
            if cell_text:
                ws.cell(row=row_counter, column=1).value = cell_text.strip()
                row_counter += 1
                cell_text = ''
            cell_text += line + '\n'
        else:
            cell_text += line + '\n'

if cell_text:
    ws.cell(row=row_counter, column=1).value = cell_text.strip()

# Save the workbook
wb.save('output.xlsx')
